"""
Professional SEO Audit Tool v3.0
By M Zahidul Islam | SEO & Search Visibility Specialist
─────────────────────────────────────────────────────────
Output Format: Matches SareeMela_COMPLETE_FINAL_v3.xlsx exactly
  Sheet 1: Complete Audit – All Pages (sectioned by page type)
  Sheet 2: Priority Action Plan

pip install streamlit pandas requests beautifulsoup4 thefuzz
            lxml openpyxl anthropic python-Levenshtein
"""

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from thefuzz import fuzz
from itertools import combinations
from concurrent.futures import ThreadPoolExecutor, as_completed
import io, json, time, logging, re
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

IGNORE_SITEMAP_KW  = ["image", "video", "attachment", "media", "gallery"]
IGNORE_EXTS        = [".jpg",".jpeg",".png",".gif",".pdf",".webp",".svg",".mp4",".mp3"]
HEADERS            = {"User-Agent": "Mozilla/5.0 (SEO-Auditor/3.0)"}

# ─────────────────────────────────────────────────────────────────
# 1. SITEMAP — iterative, no recursion
# ─────────────────────────────────────────────────────────────────
def get_sitemap_urls(sitemap_url: str) -> list[str]:
    visited, queue, found = set(), [sitemap_url], set()
    while queue:
        cur = queue.pop(0)
        if cur in visited: continue
        visited.add(cur)
        try:
            r = requests.get(cur, headers=HEADERS, timeout=15)
            r.raise_for_status()
            soup = BeautifulSoup(r.content, "xml")
            for loc in soup.find_all("loc"):
                link = loc.text.strip()
                if link.endswith(".xml"):
                    if not any(kw in link.lower() for kw in IGNORE_SITEMAP_KW):
                        queue.append(link)
                elif not any(link.lower().endswith(e) for e in IGNORE_EXTS):
                    found.add(link)
        except Exception as e:
            st.warning(f"⚠️ Sitemap error: `{cur}` — {e}")
    return list(found)


# ─────────────────────────────────────────────────────────────────
# 2. PAGE TYPE DETECTION
# ─────────────────────────────────────────────────────────────────
def detect_page_type(url: str) -> str:
    u = url.lower().rstrip("/")
    if u.count("/") <= 3: return "Homepage" if u.count("/") == 2 else "Static"
    if any(p in u for p in ["/product/", "/products/"]): return "Product"
    if any(p in u for p in ["/product-category/", "/category/", "/shop/", "/collection/"]): return "Category"
    if any(p in u for p in ["/blog/", "/post/", "/news/", "/article/"]): return "Blog"
    # WooCommerce nested category
    if u.count("/") >= 5 and "product" not in u: return "Sub-Cat"
    return "Static"


# ─────────────────────────────────────────────────────────────────
# 3. PRIMARY KEYWORD — from URL slug + Title combined
# ─────────────────────────────────────────────────────────────────
STOP_WORDS = {
    "a","an","the","and","or","but","in","on","at","to","for","of","with",
    "by","from","is","are","was","were","be","been","being","have","has",
    "had","do","does","did","will","would","could","should","may","might",
    "this","that","these","those","it","its","my","your","our","their",
    "what","which","who","how","when","where","why","all","any","both",
    "com","www","http","https","html","php","asp",
    # Bengali common stop words in English transliteration
    "er","te","ke","ba","o","e","i","ar",
}

def extract_primary_keyword(url: str, title: str) -> str:
    # From URL slug
    slug = url.rstrip("/").split("/")[-1]
    slug_words = re.sub(r"[-_]", " ", slug).lower().split()
    slug_kw = " ".join(w for w in slug_words if w not in STOP_WORDS and len(w) > 2)

    # From Title — remove site name suffix (e.g. "| SareeMela")
    clean_title = re.split(r"\s*[\|–\-]\s*[A-Z]", title)[0].strip() if title else ""
    title_words = clean_title.lower().split()
    title_kw = " ".join(w for w in title_words if w not in STOP_WORDS and len(w) > 2)

    # Prefer title if longer and more meaningful; fall back to slug
    if title_kw and len(title_kw) > len(slug_kw):
        return title_kw[:80]
    return slug_kw[:80] if slug_kw else clean_title[:80]


# ─────────────────────────────────────────────────────────────────
# 4. RULE-BASED ISSUE + FIX ENGINE
# ─────────────────────────────────────────────────────────────────
def rule_based_analysis(page: dict, cannibal_info: dict) -> tuple[str, str, str, str]:
    """
    Returns: (severity, issue, recommended_fix, priority)
    """
    issues   = []
    fixes    = []
    severity = "🟢 OK"
    priority = "🟢 LOW/OK"

    # ── Cannibalization ──────────────────────────────────────
    if cannibal_info.get("score", 0) >= 95:
        issues.append(f"Exact title duplicate with: {cannibal_info['url']}")
        fixes.append(f"301 Redirect weaker page → {cannibal_info['url']}")
        severity = "🔴 CRITICAL"; priority = "🔴 P1 — Today"
    elif cannibal_info.get("score", 0) >= 85:
        issues.append(f"High title overlap ({cannibal_info['score']}%) with: {cannibal_info['url']}")
        fixes.append("Rewrite title to target distinct search intent")
        severity = "🟠 HIGH"; priority = "🟠 P2 — This Week"
    elif cannibal_info.get("score", 0) >= 75:
        issues.append(f"Moderate title overlap ({cannibal_info['score']}%) with: {cannibal_info['url']}")
        fixes.append("Differentiate content focus and update meta title")
        severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── Status Code ──────────────────────────────────────────
    sc = page.get("status_code", 200)
    if sc == 404:
        issues.append("404 Not Found — Page is dead")
        fixes.append("Either restore the page or 301 redirect to relevant live URL")
        severity = "🔴 CRITICAL"; priority = "🔴 P1 — Today"
    elif sc in [301, 302]:
        issues.append(f"{sc} Redirect detected")
        fixes.append("Update all internal links to point directly to final destination URL")
        if severity == "🟢 OK": severity = "🟠 HIGH"; priority = "🟠 P2 — This Week"

    # ── Title ────────────────────────────────────────────────
    title = page.get("title", "")
    tlen  = len(title)
    if not title or title == "Missing":
        issues.append("Missing <title> tag")
        fixes.append("Write a unique, keyword-rich title (50–60 chars)")
        if severity == "🟢 OK": severity = "🔴 CRITICAL"; priority = "🔴 P1 — Today"
    elif tlen > 60:
        issues.append(f"Title too long ({tlen} chars, max 60)")
        fixes.append(f"Shorten title to under 60 chars. Keep primary keyword at start.")
        if severity == "🟢 OK": severity = "🟠 HIGH"; priority = "🟠 P2 — This Week"
    elif tlen < 30:
        issues.append(f"Title too short ({tlen} chars, min 30)")
        fixes.append("Expand title with supporting keyword or location modifier")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── Meta Description ─────────────────────────────────────
    md     = page.get("meta_desc", "")
    mdlen  = len(md)
    if not md or md == "Missing":
        issues.append("Missing meta description")
        fixes.append("Write a compelling 120–155 char meta description with CTA")
        if severity == "🟢 OK": severity = "🟠 HIGH"; priority = "🟠 P2 — This Week"
    elif mdlen > 160:
        issues.append(f"Meta description too long ({mdlen} chars)")
        fixes.append("Trim to under 160 chars. Keep key selling point in first 120.")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── H1 ───────────────────────────────────────────────────
    h1_count = page.get("h1_count", 0)
    if h1_count == 0:
        issues.append("Missing H1 tag")
        fixes.append("Add a single H1 containing the primary keyword")
        if severity == "🟢 OK": severity = "🟠 HIGH"; priority = "🟠 P2 — This Week"
    elif h1_count > 1:
        issues.append(f"Multiple H1 tags found ({h1_count})")
        fixes.append("Keep only one H1. Convert extras to H2 or H3.")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── Canonical ────────────────────────────────────────────
    canonical = page.get("canonical", "")
    if not canonical:
        issues.append("No canonical tag")
        fixes.append("Add self-referencing canonical tag to prevent duplicate content")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── Noindex ──────────────────────────────────────────────
    robots = page.get("robots_meta", "index")
    if "noindex" in robots.lower():
        issues.append("⚠️ Noindex meta robots — page excluded from Google")
        fixes.append("Remove noindex directive unless page is intentionally excluded")
        severity = "🔴 CRITICAL"; priority = "🔴 P1 — Today"

    # ── Schema ───────────────────────────────────────────────
    schema = page.get("schema_types", "None")
    page_type = page.get("page_type", "")
    if schema == "None":
        if page_type == "Product":
            issues.append("No Schema Markup (Product schema missing)")
            fixes.append("Add Product schema with name, price, availability, review")
        elif page_type in ["Category", "Sub-Cat"]:
            issues.append("No Schema Markup (BreadcrumbList missing)")
            fixes.append("Add BreadcrumbList schema for category navigation")
        else:
            issues.append("No Schema Markup")
            fixes.append("Add relevant schema (WebPage, Article, BreadcrumbList)")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── Thin Content ─────────────────────────────────────────
    words = page.get("word_count", 0)
    if words < 300 and page_type not in ["Homepage"]:
        issues.append(f"Thin content ({words} words, min 300)")
        fixes.append("Expand content with product details, FAQs, or buying guide")
        if severity == "🟢 OK": severity = "🟡 MEDIUM"; priority = "🟡 P3 — This Month"

    # ── OG Image ─────────────────────────────────────────────
    og_img = page.get("og_image", "Missing")
    if og_img == "Missing":
        issues.append("Missing OG Image (social sharing broken)")
        fixes.append("Add og:image meta tag with a high-quality product/category image")
        if severity == "🟢 OK": severity = "🔵 INFO"; priority = "🟢 P4 — Ongoing"

    # ── All clear ────────────────────────────────────────────
    if not issues:
        return "🟢 OK", "No issues detected", "Keep current optimization. Monitor rankings.", "🟢 LOW/OK"

    return severity, " | ".join(issues), " | ".join(fixes), priority


# ─────────────────────────────────────────────────────────────────
# 5. AI-ENHANCED FIX (Claude API — optional)
# ─────────────────────────────────────────────────────────────────
def ai_fix(page: dict, rule_issue: str, rule_fix: str, api_key: str) -> str:
    """Enhances rule-based fix with AI-generated specific recommendation."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        prompt = f"""You are an expert SEO specialist. Given this page data, write ONE concise recommended fix (max 2 sentences, actionable, specific).

URL: {page.get('url', '')}
Page Type: {page.get('page_type', '')}
Title: {page.get('title', '')}
Primary Keyword: {page.get('primary_keyword', '')}
Issues Found: {rule_issue}
Rule-based Fix: {rule_fix}

Respond with ONLY the improved recommended fix. No preamble."""

        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        return msg.content[0].text.strip()
    except Exception as e:
        logger.warning(f"AI fix failed: {e}")
        return rule_fix


# ─────────────────────────────────────────────────────────────────
# 6. PAGE SCRAPER
# ─────────────────────────────────────────────────────────────────
def scrape_page(url: str) -> dict:
    result = {
        "url": url, "page_type": detect_page_type(url),
        "status_code": 0, "title": "Missing", "title_length": 0,
        "meta_desc": "Missing", "meta_desc_length": 0,
        "h1": "Missing", "h1_count": 0,
        "canonical": "", "robots_meta": "index",
        "word_count": 0, "schema_types": "None",
        "og_image": "Missing", "response_ms": 0,
    }
    try:
        t0 = time.time()
        res = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)
        result["response_ms"]  = round((time.time() - t0) * 1000)
        result["status_code"]  = res.status_code
        if res.status_code != 200:
            return result

        soup = BeautifulSoup(res.text, "html.parser")

        # Title
        tt = soup.find("title")
        title = tt.string.strip() if tt and tt.string else ""
        result["title"]        = title or "Missing"
        result["title_length"] = len(title)

        # Meta desc
        mt = soup.find("meta", attrs={"name": "description"})
        if mt:
            md = mt.get("content", "").strip()
            result["meta_desc"]        = md or "Missing"
            result["meta_desc_length"] = len(md)

        # H1
        h1s = soup.find_all("h1")
        result["h1_count"] = len(h1s)
        result["h1"]       = h1s[0].get_text(strip=True) if h1s else "Missing"

        # Canonical
        ct = soup.find("link", rel="canonical")
        result["canonical"] = ct.get("href", "").strip() if ct else ""

        # Robots
        rt = soup.find("meta", attrs={"name": "robots"})
        if rt: result["robots_meta"] = rt.get("content", "index").lower()

        # Word count
        body = soup.find("body")
        if body: result["word_count"] = len(body.get_text(" ", strip=True).split())

        # Schema
        scripts = soup.find_all("script", type="application/ld+json")
        types = []
        for s in scripts:
            try:
                d = json.loads(s.string or "")
                if isinstance(d, list):
                    types += [x.get("@type","") for x in d if isinstance(x, dict)]
                elif isinstance(d, dict):
                    t = d.get("@type","")
                    types += t if isinstance(t, list) else ([t] if t else [])
            except: pass
        result["schema_types"] = ", ".join(types) if types else "None"

        # OG image
        og = soup.find("meta", property="og:image")
        result["og_image"] = og["content"].strip() if og and og.get("content") else "Missing"

        # Primary keyword
        result["primary_keyword"] = extract_primary_keyword(url, result["title"])

    except requests.exceptions.Timeout:
        result["status_code"] = "Timeout"
    except Exception as e:
        result["status_code"] = f"Error: {str(e)[:40]}"

    if "primary_keyword" not in result:
        result["primary_keyword"] = extract_primary_keyword(url, result.get("title",""))
    return result


# ─────────────────────────────────────────────────────────────────
# 7. PARALLEL SCRAPER
# ─────────────────────────────────────────────────────────────────
def scrape_all(urls: list[str], workers: int = 10) -> list[dict]:
    results = [None] * len(urls)
    bar = st.progress(0)
    status = st.empty()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        fmap = {ex.submit(scrape_page, u): i for i, u in enumerate(urls)}
        done = 0
        for f in as_completed(fmap):
            idx = fmap[f]
            try: results[idx] = f.result()
            except Exception as e:
                results[idx] = {"url": urls[idx], "status_code": f"Error", "title": "Error",
                                 "page_type": detect_page_type(urls[idx]),
                                 "primary_keyword": extract_primary_keyword(urls[idx], "")}
            done += 1
            bar.progress(done / len(urls))
            status.caption(f"Scanning {done}/{len(urls)} pages…")
    status.empty()
    return [r for r in results if r]


# ─────────────────────────────────────────────────────────────────
# 8. CANNIBALIZATION DETECTION
# ─────────────────────────────────────────────────────────────────
def detect_cannibalization(pages: list[dict]) -> dict[int, dict]:
    """Returns {index: {url, score}} for each page that has a conflict."""
    n = len(pages)
    best = {}  # index → {url, score}
    for i, j in combinations(range(n), 2):
        ti = str(pages[i].get("title",""))
        tj = str(pages[j].get("title",""))
        score = fuzz.token_sort_ratio(ti, tj)
        if score > 75:
            if score > best.get(i, {}).get("score", 0):
                best[i] = {"url": pages[j]["url"], "score": score}
            if score > best.get(j, {}).get("score", 0):
                best[j] = {"url": pages[i]["url"], "score": score}
    return best


# ─────────────────────────────────────────────────────────────────
# 9. AI PRIORITY ACTION PLAN (whole-site summary)
# ─────────────────────────────────────────────────────────────────
def generate_priority_plan_ai(audit_rows: list[dict], api_key: str) -> list[dict]:
    """Ask Claude to generate the Priority Action Plan sheet rows."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        critical = [r for r in audit_rows if "CRITICAL" in r.get("Severity","") or "P1" in r.get("Priority","")][:10]
        high     = [r for r in audit_rows if "HIGH" in r.get("Severity","") or "P2" in r.get("Priority","")][:10]

        summary_lines = []
        for r in critical + high:
            summary_lines.append(
                f"- [{r['Page Type']}] {r['URL']}\n"
                f"  Issue: {r['Issue / Problem']}\n"
                f"  Current fix: {r['Recommended Fix']}"
            )

        prompt = f"""You are a senior SEO strategist. Based on these audit findings, generate a prioritized action plan.

AUDIT FINDINGS:
{chr(10).join(summary_lines)}

Return a JSON array of action items. Each item must have exactly these keys:
- "Priority": one of "🔴 P1 — Today", "🟠 P2 — This Week", "🟡 P3 — This Month", "🟢 P4 — Ongoing"
- "Action Required": specific SEO action (1 sentence)
- "URL(s) Affected": the URL or "Multiple pages"
- "Expected Impact": business/SEO impact (short phrase)

Return ONLY valid JSON array. No explanation. No markdown."""

        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```json|^```|```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        logger.warning(f"AI action plan failed: {e}")
        return []


# ─────────────────────────────────────────────────────────────────
# 10. RULE-BASED PRIORITY ACTION PLAN
# ─────────────────────────────────────────────────────────────────
def build_rule_action_plan(audit_rows: list[dict]) -> list[dict]:
    plan = []
    priority_order = ["🔴 P1 — Today", "🟠 P2 — This Week", "🟡 P3 — This Month", "🟢 P4 — Ongoing"]
    for row in audit_rows:
        if row.get("Severity") in ["🟢 OK", "🔵 INFO"]:
            continue
        plan.append({
            "Priority":        row.get("Priority", "🟢 P4 — Ongoing"),
            "Action Required": row.get("Recommended Fix", "")[:200],
            "URL(s) Affected": row.get("URL",""),
            "Expected Impact": _impact_for(row.get("Issue / Problem", ""), row.get("Page Type",""))
        })
    plan.sort(key=lambda x: priority_order.index(x["Priority"]) if x["Priority"] in priority_order else 99)
    return plan


def _impact_for(issue: str, page_type: str) -> str:
    issue_l = issue.lower()
    if "404"       in issue_l: return "Recovers crawl budget & lost backlinks"
    if "noindex"   in issue_l: return "Restores Google indexing for this page"
    if "duplicate" in issue_l or "cannibal" in issue_l:
        return "Consolidates ranking signals, improves position"
    if "title"     in issue_l: return "Improves CTR in search results"
    if "meta desc" in issue_l: return "Boosts click-through rate from SERPs"
    if "h1"        in issue_l: return "Strengthens on-page keyword relevance"
    if "schema"    in issue_l: return "Enables rich snippets in search results"
    if "thin"      in issue_l: return "Improves content quality score & rankings"
    if "canonical" in issue_l: return "Prevents duplicate content dilution"
    if "og image"  in issue_l: return "Improves social media click-through"
    return "Improves overall SEO health"


# ─────────────────────────────────────────────────────────────────
# 11. EXCEL REPORT — SareeMela format exactly
# ─────────────────────────────────────────────────────────────────
SEVERITY_COLORS = {
    "🔴 CRITICAL": "FFCDD2",
    "🟠 HIGH":     "FFE0B2",
    "🟡 MEDIUM":   "FFF9C4",
    "🟢 OK":       "E8F5E9",
    "🔵 INFO":     "E3F2FD",
}
PRIORITY_COLORS = {
    "🔴 P1 — Today":       "FFCDD2",
    "🟠 P2 — This Week":   "FFE0B2",
    "🟡 P3 — This Month":  "FFF9C4",
    "🟢 P4 — Ongoing":     "E8F5E9",
    "🟢 LOW/OK":           "E8F5E9",
}
SECTION_ORDER = ["Homepage", "Static", "Category", "Sub-Cat", "Blog", "Product"]

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_excel_report(audit_rows: list[dict], action_plan: list[dict], site_url: str) -> bytes:
    wb = openpyxl.Workbook()

    # ── SHEET 1: Complete Audit ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Complete Audit – All Pages"

    # Big title row
    site_name = re.sub(r"https?://", "", site_url).split("/")[0]
    from datetime import datetime
    month_year = datetime.now().strftime("%B %Y")

    ws1.merge_cells("A1:I1")
    ws1["A1"] = (
        f"{site_name} — COMPLETE Keyword Cannibalization Report  |  "
        f"Every Single Page Audited  |  "
        f"Static + Category + Blog + Products  |  {month_year}"
    )
    ws1["A1"].font      = Font(bold=True, size=11, color="FFFFFF")
    ws1["A1"].fill      = _fill("1A237E")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[1].height = 30

    # Legend row
    ws1.merge_cells("A2:I2")
    ws1["A2"] = (
        "🔴 CRITICAL = Fix Today  |  🟠 HIGH = This Week  |  "
        "🟡 MEDIUM = This Month  |  🟢 LOW/OK = Monitor/No Action  |  🔵 INFO = No Cannibalization Risk"
    )
    ws1["A2"].font      = Font(italic=True, size=9, color="333333")
    ws1["A2"].fill      = _fill("F5F5F5")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[2].height = 20

    # Column headers
    col_headers = ["#", "Page Type", "Severity", "URL",
                   "Page Title (Fetched Live)", "Primary Keyword",
                   "Issue / Problem", "Recommended Fix", "Priority"]
    col_widths   = [5, 10, 13, 45, 40, 28, 40, 45, 18]

    ws1.append(col_headers)
    header_row = ws1.max_row
    for col_idx, (hdr, width) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws1.cell(row=header_row, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _fill("283593")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[header_row].height = 22

    # Group rows by section
    grouped = {s: [] for s in SECTION_ORDER}
    for row in audit_rows:
        pt = row.get("Page Type", "Static")
        grouped.setdefault(pt, []).append(row)

    row_num = header_row + 1
    counter = 1

    for section in SECTION_ORDER:
        rows = grouped.get(section, [])
        if not rows: continue

        # Section divider
        ws1.merge_cells(f"A{row_num}:I{row_num}")
        label = f"━━━  SECTION — {section.upper()} PAGES  ({len(rows)} Pages)  ━━━"
        cell = ws1.cell(row=row_num, column=1, value=label)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = _fill("37474F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[row_num].height = 18
        row_num += 1

        for r in rows:
            sev   = r.get("Severity", "🟢 OK")
            color = SEVERITY_COLORS.get(sev, "FFFFFF")
            vals  = [
                counter,
                r.get("Page Type",""),
                sev,
                r.get("URL",""),
                r.get("Page Title",""),
                r.get("Primary Keyword",""),
                r.get("Issue / Problem",""),
                r.get("Recommended Fix",""),
                r.get("Priority",""),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws1.cell(row=row_num, column=col_idx, value=val)
                cell.fill      = _fill(color)
                cell.border    = _border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx == 4:   # URL — make it a hyperlink style
                    cell.font = Font(color="1565C0", underline="single", size=9)
                elif col_idx in [1, 2, 3, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.font = Font(size=9)
                else:
                    cell.font = Font(size=9)
            ws1.row_dimensions[row_num].height = 40
            row_num += 1
            counter += 1

    ws1.freeze_panes = "A4"

    # ── SHEET 2: Priority Action Plan ─────────────────────────
    ws2 = wb.create_sheet("Priority Action Plan")

    site_name2 = re.sub(r"https?://", "", site_url).split("/")[0]
    ws2.merge_cells("A1:D1")
    ws2["A1"] = (
        f"{site_name2} — Priority Action Plan  |  "
        "P1 Fix Today → P2 This Week → P3 This Month → P4 Ongoing"
    )
    ws2["A1"].font      = Font(bold=True, size=11, color="FFFFFF")
    ws2["A1"].fill      = _fill("1A237E")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    plan_headers = ["Priority", "Action Required", "URL(s) Affected", "Expected Impact"]
    plan_widths  = [22, 60, 50, 38]
    ws2.append(plan_headers)
    hrow = ws2.max_row
    for col_idx, (hdr, w) in enumerate(zip(plan_headers, plan_widths), 1):
        cell = ws2.cell(row=hrow, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _fill("283593")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[hrow].height = 20

    for item in action_plan:
        pri   = item.get("Priority","🟢 P4 — Ongoing")
        color = PRIORITY_COLORS.get(pri, "FFFFFF")
        vals  = [
            pri,
            item.get("Action Required",""),
            item.get("URL(s) Affected",""),
            item.get("Expected Impact",""),
        ]
        rn = ws2.max_row + 1
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=rn, column=col_idx, value=val)
            cell.fill      = _fill(color)
            cell.border    = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font      = Font(size=9, bold=(col_idx==1))
        ws2.row_dimensions[rn].height = 35

    ws2.freeze_panes = "A3"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────────
# 12. STREAMLIT UI
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SEO Audit Tool v3 | Zahidul Islam",
    layout="wide", page_icon="🛡️"
)

st.markdown("""
<div style='background:linear-gradient(135deg,#1a237e,#283593,#0d47a1);
     padding:1.8rem 2rem;border-radius:12px;margin-bottom:1.5rem'>
  <h1 style='color:#fff;margin:0;font-size:1.9rem'>🛡️ Professional SEO Audit Tool <span style="font-size:1rem;opacity:.7">v3.0</span></h1>
  <p style='color:#90caf9;margin:.4rem 0 0'>
    Deep On-Page · Schema · Cannibalization · AI Action Plan &nbsp;|&nbsp;
    Output matches SareeMela-style professional Excel report
  </p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")
    sitemap_input = st.text_input("Sitemap URL", placeholder="https://example.com/sitemap_index.xml")
    max_urls      = st.number_input("Max Pages", 50, 2000, 500, 50)
    workers       = st.slider("Parallel Requests", 3, 20, 8)

    st.divider()
    st.subheader("🤖 AI Enhancement")
    st.caption("Optional — adds Claude AI to write smarter fixes & action plan.")
    api_key    = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-…")
    use_ai_fix = st.checkbox("AI-powered Recommended Fix (per page)", value=False,
                              help="Slower — makes 1 API call per page with issues")
    use_ai_plan = st.checkbox("AI-generated Priority Action Plan", value=bool(api_key))

    st.divider()
    run = st.button("🚀 Run Audit", type="primary", use_container_width=True)

# ── Main ───────────────────────────────────────────────────────
if run:
    if not sitemap_input.strip():
        st.error("Please enter a sitemap URL.")
        st.stop()

    # Step 1 — Sitemap
    with st.spinner("📡 Fetching sitemap…"):
        all_links = get_sitemap_urls(sitemap_input.strip())
    if not all_links:
        st.error("❌ No pages found. Check your sitemap URL.")
        st.stop()

    urls = all_links[:max_urls]
    st.success(f"✅ Found **{len(all_links)}** pages. Auditing **{len(urls)}**.")

    # Step 2 — Scrape
    st.subheader("⏳ Scanning pages…")
    pages = scrape_all(urls, workers)

    # Step 3 — Cannibalization
    with st.spinner("🔍 Detecting cannibalization…"):
        cannibal_map = detect_cannibalization(pages)

    # Step 4 — Build audit rows
    audit_rows = []
    ai_fix_bar = None
    pages_with_issues = [p for i,p in enumerate(pages)
                         if cannibal_map.get(i) or p.get("status_code") != 200
                         or not p.get("title") or p.get("title")=="Missing"]

    if use_ai_fix and api_key and pages_with_issues:
        st.info(f"🤖 AI is writing fixes for {len(pages_with_issues)} pages with issues…")

    for i, page in enumerate(pages):
        cannibal_info = cannibal_map.get(i, {})
        severity, issue, fix, priority = rule_based_analysis(page, cannibal_info)

        if use_ai_fix and api_key and severity not in ["🟢 OK", "🔵 INFO"]:
            fix = ai_fix(page, issue, fix, api_key)

        audit_rows.append({
            "#":                  i + 1,
            "Page Type":          page.get("page_type",""),
            "Severity":           severity,
            "URL":                page.get("url",""),
            "Page Title":         page.get("title",""),
            "Primary Keyword":    page.get("primary_keyword",""),
            "Issue / Problem":    issue,
            "Recommended Fix":    fix,
            "Priority":           priority,
            # extra for action plan
            "_status_code":       page.get("status_code",""),
            "_cannibal_url":      cannibal_info.get("url",""),
            "_cannibal_score":    cannibal_info.get("score",0),
        })

    # Step 5 — Action Plan
    if use_ai_plan and api_key:
        with st.spinner("🤖 Claude is writing your Priority Action Plan…"):
            action_plan = generate_priority_plan_ai(audit_rows, api_key)
        if not action_plan:
            action_plan = build_rule_action_plan(audit_rows)
    else:
        action_plan = build_rule_action_plan(audit_rows)

    # Step 6 — Summary metrics
    st.subheader("📊 Audit Overview")
    sev_counts = pd.Series([r["Severity"] for r in audit_rows]).value_counts()
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Total Pages",     len(audit_rows))
    m2.metric("🔴 Critical",     sev_counts.get("🔴 CRITICAL",0))
    m3.metric("🟠 High",         sev_counts.get("🟠 HIGH",0))
    m4.metric("🟡 Medium",       sev_counts.get("🟡 MEDIUM",0))
    m5.metric("🟢 OK",           sev_counts.get("🟢 OK",0))

    # Step 7 — Preview tables
    df_audit  = pd.DataFrame(audit_rows)
    df_plan   = pd.DataFrame(action_plan)
    display_cols = ["#","Page Type","Severity","URL","Page Title",
                    "Primary Keyword","Issue / Problem","Recommended Fix","Priority"]

    tab1, tab2, tab3 = st.tabs(["📋 Full Audit", "🚨 Issues Only", "📌 Action Plan"])
    with tab1:
        st.dataframe(df_audit[display_cols], use_container_width=True, height=500)
    with tab2:
        issues_only = df_audit[df_audit["Severity"] != "🟢 OK"]
        st.info(f"{len(issues_only)} pages need attention.")
        st.dataframe(issues_only[display_cols], use_container_width=True, height=500)
    with tab3:
        if not df_plan.empty:
            st.dataframe(df_plan, use_container_width=True, height=400)
        else:
            st.info("No action plan items generated.")

    # Step 8 — Export
    st.subheader("📥 Download Report")
    with st.spinner("Building Excel report…"):
        excel_bytes = build_excel_report(audit_rows, action_plan, sitemap_input.strip())

    site_slug = re.sub(r"https?://", "", sitemap_input).split("/")[0].replace(".", "_")
    st.download_button(
        label     = "📥 Download Professional Excel Report",
        data      = excel_bytes,
        file_name = f"SEO_Audit_{site_slug}.xlsx",
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ── Footer ─────────────────────────────────────────────────────
st.markdown("---")
st.caption("© M Zahidul Islam | SEO & Search Visibility Specialist | v3.0")
